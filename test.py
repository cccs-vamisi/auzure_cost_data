import os
import json
import asyncio
from azure.core.exceptions import HttpResponseError
from azure.identity import DefaultAzureCredential
from azure.mgmt.costmanagement import CostManagementClient
from azure.mgmt.costmanagement.models import QueryDefinition, QueryDataset, QueryTimePeriod, QueryAggregation, QueryGrouping
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook, load_workbook
from dateutil.relativedelta import relativedelta 
import time





def file_reading():
    work_book = Workbook()
    work_sheet = work_book.active
    files_to_read = ["Tenant", "Resource Group", "Team", "CBR" ]
    resource_group_list = []
    subscription_Id = []
    cell = ["A", "B", "C", "D" ]
    
    # Read text files and populate data inside a python list
    for i in range(0, len(files_to_read)):
        list_data = []
        f = open(f"{files_to_read[i]}.txt", "r")

        for x in f:
            x = x.replace("\n", "")
            if len(x) == 0:
                continue
            else:
                if f"{files_to_read[i]}.txt" == "Resource Group.txt":
                    resource_group_list.append(x)
                list_data.append(x)
        f.close()
        
        # Read python list and populate data inside an excel sheet
        work_sheet[f'{cell[i]}{1}'] = files_to_read[i]
        start_row = 2
        start_column = i + 1
        
        for i, value in enumerate(list_data):
            work_sheet.cell(row=start_row+i, column=start_column, value=value)
    
    # Populate python list with subscription IDs based on the resource group list
    f = open(f"sub_id.txt", "r")
    for x in f:
        x = x.replace("\n", "")
        if len(x) == 0:
            continue
        else:
            subscription_Id.append(x)
    f.close()
    
    # Create a loop that will use the azure api to get data for the last 12 month based on the 
    

    # work_book.save("APA_Cost_History.xlsx")
file_reading()