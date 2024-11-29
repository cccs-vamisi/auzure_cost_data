
import os
import asyncio
from azure.core.exceptions import HttpResponseError
from azure.identity import DefaultAzureCredential
from azure.mgmt.costmanagement import CostManagementClient
from azure.mgmt.costmanagement.models import QueryDefinition, QueryDataset, QueryTimePeriod, QueryAggregation
from datetime import datetime, timedelta
from openpyxl import Workbook
from dateutil.relativedelta import relativedelta 
import time

# Initialize the Azure Cost Management client with authentication:
credential = DefaultAzureCredential()
client = CostManagementClient(credential)

current_date =  datetime.now()
# The following code will calculate the time frame for the last 12 months
start_date_of_last_twelve_months = (current_date.replace(day=1) - relativedelta(months=1)).replace(day=1) - relativedelta(months=11)
end_date_of_last_twelve_months = current_date.replace(day=1) - timedelta(days=1)

# This class will contain the methods to get the data
class Data():
    def __init__(self, resource_group,
                 subscription_id,
                 scope,
                 final_returned_data=dict()):
        self.resource_group = resource_group
        self.subscription_id = subscription_id
        self.scope = scope
        self.final_returned_data = final_returned_data
        
    # This helper method will help remove repetitive code
    def helper_method(self, from_date, to_date, scope_value):
        time_period=QueryTimePeriod(from_property=from_date, to=to_date)
        query_aggregation = dict()
        query_aggregation["totalCost"] = QueryAggregation(name="Cost", function="Sum")
        querydataset = QueryDataset(granularity="Monthly", configuration=None, aggregation=query_aggregation)
        query = QueryDefinition(type="ActualCost", timeframe="Custom", time_period=time_period, dataset=querydataset)
        while True:
            try:
                    # Make the API request
                    result = client.query.usage(scope=scope_value, parameters=query)
                    return result.rows
            except HttpResponseError as e:
                    # Check for rate-limiting headers
                    retry_after_qpu = int(e.response.headers.get('x-ms-ratelimit-microsoft.costmanagement-qpu-retry-after', 0))
                    retry_after_entity = int(e.response.headers.get('x-ms-ratelimit-microsoft.costmanagement-entity-retry-after', 0))
                    retry_after_tenant = int(e.response.headers.get('x-ms-ratelimit-microsoft.costmanagement-tenant-retry-after', 0))
                    retry_after_client = int(e.response.headers.get('x-ms-ratelimit-microsoft.costmanagement-client-retry-after', 0))
                    # Determine the maximum retry time
                    max_retry_after = max(retry_after_qpu, retry_after_entity, retry_after_tenant, retry_after_client)
                    if max_retry_after > 0:
                        print(f"Rate-limited. Retrying in {max_retry_after} seconds...")
                        time.sleep(max_retry_after)
                    else:
                        raise e  # Re-raise the exception if no retry headers are present
    async def fetch_cost_for_last_twelve_months(self):
        return self.helper_method(start_date_of_last_twelve_months, end_date_of_last_twelve_months, self.scope)
# resource_group, subscription_id

async def main_method(resource_group, subscription_id):
    scope = f'/subscriptions/{subscription_id}/resourceGroups/{resource_group}'
    obj = Data(resource_group, subscription_id, scope)
    task1 = asyncio.create_task(obj.fetch_cost_for_last_twelve_months())
    obj.final_returned_data["last_twelve_months"] = await task1
    return obj.final_returned_data


def format_data(data):
    list_item = data["last_twelve_months"]
    for i in range(0, len(list_item)):
        date_str = list_item[i][1]
        date_obj = datetime.fromisoformat(date_str)
        month = date_obj.strftime("%B")
        year =  date_obj.strftime("%Y")
        list_item[i][1] = f"{month}, {year}"
    return data

def file_reading():
    work_book = Workbook()
    work_sheet = work_book.active
    files_to_read = ["Tenant", "Resource Group", "Team", "CBR" ]
    resource_group_list = []
    subscription_Id_list = []
    cell = ["A", "B", "C", "D" ]
    
    # Read text files and populate data inside a python list
    for i in range(0, len(files_to_read)):
        list_data = []
        f = open(f"{files_to_read[i]}.txt", "r")

        for x in f:
            x = x.replace("\n", "")
            if len(x) == 0:
                continue
            else:
                if f"{files_to_read[i]}.txt" == "Resource Group.txt":
                    resource_group_list.append(x)
                list_data.append(x)
        f.close()
        
        # Read python list and populate data inside an excel sheet
        work_sheet[f'{cell[i]}{1}'] = files_to_read[i]
        start_row = 2
        start_column = i + 1
        
        for i, value in enumerate(list_data):
            work_sheet.cell(row=start_row+i, column=start_column, value=value)
    
    # Populate python list with subscription IDs based on the resource group list
    f = open(f"sub_id.txt", "r")
    for x in f:
        x = x.replace("\n", "")
        if len(x) == 0:
            continue
        else:
            subscription_Id_list.append(x)
    f.close()
    
    # Create a loop that will use the azure api to get data for the last 12 month based on the 
    for i in range(0, len(subscription_Id_list)):
        date_populated = False
        resource_group = resource_group_list[i]
        subscription_id = subscription_Id_list[i]
        unformatted_data = asyncio.run(main_method(resource_group, subscription_id))
        formatted_data = format_data(unformatted_data)
        formatted_data = formatted_data['last_twelve_months']
        for j in range(0, len(formatted_data)):
            start_column = j + 5
            cost_value = round(formatted_data[j][0], 2)
            if date_populated == False:
                date_value = formatted_data[j][1]
                work_sheet.cell(row=1, column=start_column, value=date_value)

            work_sheet.cell(row=2+i, column=start_column, value=cost_value)
            date_populated == True
    work_book.save("APA_Cost_History.xlsx")
    
file_reading()
    




























